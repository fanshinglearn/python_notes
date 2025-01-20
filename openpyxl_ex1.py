import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 創建新檔案
wb = Workbook()

# 讀取檔案
wb = load_workbook('test.xlsx')

# 存檔
wb.save('output.xlsx')

# 選取工作表
ws = wb.active
ws = wb['Sheet1']
print(ws)

# 創建工作表
wb.create_sheet('123')
ws.title = '123'
print(wb.sheetnames)

# 取得儲存格值
# 第一種寫法
a = ws['A1'].value
print(a)

# 第二種寫法
a = ws['A1']
print(a.value)

# 修改儲存格資料
ws['B3'] = '123'
print(ws['B3'].value)

# 新增List
ws.append([123,None,789,0])

# 顯示範圍資料
for row in range(2,6):
    for col in range(2,4):
        char = get_column_letter(col)
        print(ws[f'{char}{row}'].value)
