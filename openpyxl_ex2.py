import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('test.xlsx')
ws = wb.active

# 設定置中對齊
from openpyxl.styles import Alignment
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
# 靠左置中
ws['A1'].alignment = Alignment(vertical='center')

# 當下時間
import datetime
ws['A1'] = datetime.datetime.now()

# 超連結
ws['A1'].hyperlink = 'https://www.google.com.tw/?hl=zh_TW'

# 插入圖片
from openpyxl.drawing.image import Image
img = Image('test.png')
print(img.width)
print(img.height)
img.width /= 4
img.height /= 4
ws.add_image(img, 'A1')

# 修改儲存高寬高
ws.column_dimensions['A'].width = 100
ws.row_dimensions[1].height = 100

# ___________________________________________

# 合併儲存格
ws.merge_cells('A1:E1')
ws.unmerge_cells('A1:E1')

# 插入橫直排
ws.insert_rows(3)
ws.insert_cols(2)
ws.delete_cols(2)
ws.delete_cols(2)
ws.delete_cols(2)

# 範圍移動
ws.move_range('A1:E5', rows=2, cols=3)
